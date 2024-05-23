from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib import messages
from .models import League, Team, Player, Auction
from .utils import generate_bracket
from django import forms
from openpyxl import load_workbook
from datetime import datetime, timedelta

@login_required
def my_team(request):
    user = request.user
    team = Team.objects.filter(owner=user).first()
    leagues = League.objects.all()

    if request.method == 'POST':
        if 'rename_team' in request.POST:
            team_name = request.POST.get('team_name')
            if team_name:
                team.name = team_name
                team.save()
                messages.success(request, 'Team renamed successfully.')
        
        elif 'enter_league' in request.POST:
            league_id = request.POST.get('league_id')
            league = get_object_or_404(League, id=league_id)
            team.leagues.add(league)
            messages.success(request, 'Team entered into league successfully.')

        elif 'delete_team' in request.POST:
            team.delete()
            messages.success(request, 'Team deleted successfully.')
            return redirect('home')
        
        elif 'create_team' in request.POST:
            new_team_name = request.POST.get('new_team_name')
            if new_team_name:
                new_team = Team.objects.create(name=new_team_name, owner=user)
                messages.success(request, 'New team created successfully.')
                return redirect('my_team')

    context = {
        'team': team,
        'leagues': leagues,
    }
    return render(request, 'my_team.html', context)

@login_required
def assign_players(request):
    user = request.user
    team = Team.objects.filter(owner=user).first()
    players = Player.objects.all()

    if request.method == 'POST':
        form = AssignPlayerForm(request.POST)
        if form.is_valid():
            player_id = form.cleaned_data['player']
            player = get_object_or_404(Player, id=player_id)
            team.players.add(player)
            messages.success(request, f'Player {player.name} assigned to your team.')
            return redirect('assign_players')
    else:
        form = AssignPlayerForm()

    context = {
        'team': team,
        'players': players,
        'form': form,
    }
    return render(request, 'assign_players.html', context)

class TeamCreationForm(forms.ModelForm):
    class Meta:
        model = Team
        fields = ['name', 'leagues']

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

@login_required
def home(request):
    user = request.user
    teams = Team.objects.filter(owner=user)
    leagues = League.objects.filter(team__owner=user).distinct()  # Fetch leagues through teams

    if request.method == 'POST':
        if 'create_team' in request.POST:
            form = TeamCreationForm(request.POST)
            if form.is_valid():
                team = form.save(commit=False)
                team.owner = user
                team.save()
                messages.success(request, 'New team created successfully.')
                return redirect('home')
        else:
            form = TeamCreationForm()
    else:
        form = TeamCreationForm()

    context = {
        'form': form,
        'teams': teams,
        'leagues': leagues,
    }
    return render(request, 'home.html', context)

    context = {
        'login_form': login_form,
        'register_form': register_form,
    }
    if request.user.is_authenticated:
        user_leagues = request.user.leagues.all()
        open_leagues = League.objects.exclude(id__in=user_leagues.values_list('id', flat=True))
        context['user_leagues'] = user_leagues
        context['open_leagues'] = open_leagues

    return render(request, 'home.html', context)

@login_required
def create_league(request):
    if not request.user.is_superuser:
        messages.error(request, 'Only administrators can use this function.')
        return redirect('home')
    if request.method == 'POST':
        league_name = request.POST.get('league_name')
        if league_name:
            league = League.objects.create(name=league_name)
            league.members.add(request.user)
            return redirect('league_detail', league_id=league.id)
    return render(request, 'create_league.html')

@login_required
def league_detail(request, league_id):
    league = get_object_or_404(League, id=league_id)
    teams = league.teams.all()
    members_teams = {member: member.team_set.filter(leagues=league) for member in league.members.all()}
    
    bracket = generate_bracket(league.members.all())
    
    return render(request, 'league_detail.html', {
        'league': league,
        'teams': teams,
        'members_teams': members_teams,
        'bracket': bracket,
    })

@login_required
def team_detail(request, team_id):
    team = get_object_or_404(Team, id=team_id, owner=request.user)
    return render(request, 'team_detail.html', {'team': team})

@login_required
def create_auction(request, league_id):
    league = get_object_or_404(League, id=league_id)
    if request.user.is_superuser and request.method == 'POST':
        player_id = request.POST.get('player_id')
        player = get_object_or_404(Player, id=player_id)
        start_time = datetime.now()
        end_time = start_time + timedelta(seconds=20)
        Auction.objects.create(league=league, player=player, start_time=start_time, end_time=end_time)
    return redirect('league_detail', league_id=league_id)

class UploadFileForm(forms.Form):
    file = forms.FileField()

@login_required
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = load_workbook(file)
            sheet = wb.active
            players_created = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rank, name, team, position, bye_week = row
                Player.objects.create(rank=rank, name=name, team=team, position=position, bye_week=bye_week)
                players_created += 1
            messages.success(request, f'Successfully uploaded {players_created} players.')
            return redirect('home')
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})

def all_leagues(request):
    leagues = League.objects.all()
    return render(request, 'all_leagues.html', {'leagues': leagues})

@login_required
def all_teams(request):
    if request.method == 'POST':
        form = TeamCreationForm(request.POST)
        if form.is_valid():
            team = form.save(commit=False)
            team.owner = request.user
            team.save()
            form.save_m2m()  # Save the many-to-many relationships
            return redirect('all_teams')
    else:
        form = TeamCreationForm()

    teams = Team.objects.all()
    return render(request, 'all_teams.html', {'teams': teams, 'form': form})
