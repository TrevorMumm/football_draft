from django.contrib import admin
from .models import Player, League, Team, Auction
from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
from django.http import HttpResponseRedirect
from django.contrib import messages
from .models import Player
from .forms import UploadFileForm
from openpyxl import load_workbook

class PlayerAdmin(admin.ModelAdmin):
    change_list_template = "admin/upload_players_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload/', self.admin_site.admin_view(self.upload_players), name='upload_players'),
        ]
        return custom_urls + urls

    def upload_players(self, request):
        if request.method == 'POST':
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES['file']
                wb = load_workbook(file)
                sheet = wb.active
                players_created = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    rank, name, team, position, bye_week = row
                    Player.objects.create(rank=rank, name=name, team=team, position=position, bye_week=bye_week)
                    players_created += 1
                self.message_user(request, f'Successfully uploaded {players_created} players.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            form = UploadFileForm()
        context = {
            'form': form,
            'title': 'Upload Player Data',
        }
        return render(request, 'admin/upload_players.html', context)

admin.site.register(Player, PlayerAdmin)
admin.site.register(League)
admin.site.register(Team)
admin.site.register(Auction)
